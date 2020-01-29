import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.action_chains import ActionChains

## Excel Dosyamızın Path'ini verdik.
path= r"C:\Users\MUSTAFA\Desktop\testdata.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
print(workbook.sheetnames)
#Satır ve Kolon Sayımızı aldık
#.For'da kullanacağız.
rows = sheet.max_row
cols = sheet.max_column

## Browser açılır
## Google Search Bar elementi seçilir.
browser = webdriver.Firefox()
browser.get("https://www.google.com")
time.sleep(1)

## Tüm hücreleri sırasıyla dolaşacak for döngümüz.
## aranacak = o hücrelerdeki değerler
## searchbar = Google Arama Barı name ' e göre seçildi.
## Enter dedik ve 2 sn sayfa yüklenmesini bekledik.
## Ardından elementin içeriğini temizledik.
for r in range(1,rows+1):
    for c in range(1,cols+1):
        aranacak = sheet.cell(row=r,column=c).value
        action = ActionChains(browser)
        searchbar = browser.find_element_by_name("q")
        searchbar.send_keys(aranacak)
        action.send_keys(Keys.ENTER)
        time.sleep(2)
        action.perform()
        browser.find_element_by_name("q").clear()

